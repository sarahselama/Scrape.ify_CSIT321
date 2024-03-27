import json
import requests
from bs4 import BeautifulSoup as bs
import string
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

headers = {
    'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'
}

def product_name_faces(quary_name):
    source = 'Faces'
    print('\nScraping faces..')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'}
    input_url = quary_name.replace(' ', '+')
    url = f'https://www.faces.com/ae-en/search?q={input_url}&lang=en_AE'
 # Generate the search URL
    input_url = quary_name.replace(' ', '+')
    url = f'https://www.faces.com/ae-en/search?q={input_url}&lang=en_AE'

    # Make the request and parse the HTML
    r = requests.get(url, headers=headers)
    soup = bs(r.content, 'html.parser')

    # Find all product links
    product_links = soup.find_all('a', class_='product-tile-name-link')

    products = []
    if len(product_links) > 10:
        product_links = product_links[:10]
    # Iterate over each product link to extract details
    for link in product_links:
        product_url = f"https://www.faces.com{link['href']}"
        print(f'Scraping url: {product_url}')
        res = requests.get(product_url)
        product_soup = bs(res.content, 'html.parser')

        try:
            product_name = product_soup.find('div', class_='product-name').text.replace('\n', ' ').strip()
        except:
            product_name = ''

        try:
            product_image = product_soup.find('img', class_='d-inline-block img-fluid product-image-holder').get('src')
        except:
            product_image = ''

        try:
            product_price = product_soup.find('span', class_='js-price')['data-original-price']
        except:
            product_price = ''

        try:
            pid = str(product_soup.find('div', attrs={'id': 'yotpo-bottomline-top-div'})['data-product-id'])
            data = {
                'methods': f'[{{"method":"bottomline","params":{{"pid":"{pid}","link":"","skip_average_score":false,"main_widget_pid":"{pid}","index":1,"element_id":"2"}}}}]',
                'app_key': 'GMkxpzEQegQXCU7Kmy3kxY8N8PKI7OiI11ZxIKvi',
                'is_mobile': 'false',
                'widget_version': '2021-08-12_14-50-29'
            }
            rating_url = f'https://staticw2.yotpo.com/batch/app_key/GMkxpzEQegQXCU7Kmy3kxY8N8PKI7OiI11ZxIKvi/domain_key/{pid}/widget/bottomline'
            rating_res = requests.post(rating_url, headers=headers, json=data)
            rating_html = rating_res.json()[0]['result']
            rating_soup = bs(rating_html, 'html.parser')
            product_rating = rating_soup.find('span', {'class': 'sr-only'}).text.replace(' rating', '').replace('star', '').strip()
            total_reviews = rating_soup.find('a', attrs={'class': 'text-m'})['aria-label'].replace('reviews', '').strip()
        except:
            product_rating = ''
            total_reviews = ''

        # Add product details to the list
        datan = {
            "Product URL": product_url,
            "Quary_Searched" : quary_name,
            "Source": source,
            "Product Name": product_name,
            "Product Rating": product_rating,
            "Total Reviews": total_reviews,
            "Product Price": product_price,
            "Product Image": product_image,


        }
        products.append(datan)

    return products




# Define a function to retrieve product details based on the given product name
def product_name_sephora(quary_name):
    source = 'Sephora'
    print('\nScraping sephora..')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'}

    # Replace spaces with '+' in the product name for the URL
    input_url = quary_name.replace(' ', '+')

    # Construct the search URL for the Sephora website
    url = f'https://www.sephora.ae/en/search?q={input_url}'
    # Send a GET request to the URL using the defined headers
    r = requests.get(url, headers=headers)
    r.raise_for_status()  # Raise an exception for HTTP errors
    soup = bs(r.content, 'html.parser')

    # Attempt to find all product information containers on the page
    product_infos = soup.find_all('div', class_='product-info')

    # List to store product details
    products = []
    if len(product_infos)>10:
        product_infos = product_infos[:10]
    # Loop through each product information container found
    for product_info in product_infos:
        # Attempt to find the link element that contains the product name
        product_url = product_info.find('a', class_='product-tile-link').get('href')
        print(f'Scraping product: {product_url}')
        res = requests.get(product_url, headers=headers)
        res.raise_for_status()  # Raise an exception for HTTP errors
        soup = bs(res.content, 'html.parser')

        # Extract product details from the product page
        product_name = soup.find('span', class_='product-name product-name-bold').text.strip() if soup.find('span', class_='product-name product-name-bold') else ''
        product_image = soup.find('img', class_='primary-image primary-image-vertical-mode').get('src') if soup.find('img', class_='primary-image primary-image-vertical-mode') else ''
        product_price = soup.find('span', class_='price-sales price-sales-standard').text.replace('(1)', '').strip() if soup.find('span', class_='price-sales price-sales-standard') else ''
        product_rating = soup.find('div', class_='bv-overall-score').text.split('/5')[0].strip() if soup.find('div', class_='bv-overall-score') else ''
        total_reviews = soup.find('span', class_='bv-number-review').text.replace('reviews', '').strip() if soup.find('span', class_='bv-number-review') else ''
        datan = {
            "Product URL": product_url,
            "Quary_Searched" : quary_name,
            "Source": source,
            "Product Name": product_name,
            "Product Rating": product_rating,

            "Total Reviews": total_reviews,
            "Product Price": product_price,
            "Product Image": product_image,
        }
        # Append product details to the products list
        products.append(datan)

    return products  # Return the list of product details


def product_name_beautybay(quary_name):
  source = 'Beauty Bay'
  print('\nScraping Beauty Bay..')
  headers = {
        'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'
    }
    # Replace spaces with '%20' in the product name for the URL
  input_url = quary_name.replace(' ', '%20')
  # Construct the search URL for the Beauty Bay website
  url = f'https://lister-page-api.public.prd.beautybay.com/listings/search?pageUrl=%2Fl%2F&filters=%7B%7D&sortOption=%7B%22sortValue%22%3Anull%2C%22sortOrder%22%3Anull%2C%22value%22%3A%22null%22%7D&page=1&pageSize=72&search={input_url}&userId=279add6b-b67c-4133-a753-8c430b0dcd1b&sessionId=8f20e289-4f44-453c-9a5e-87af9d497443&preview=false&referrer=&platform=attraqt&noRestrictions=false&debug=false&locale=en-GB&bagId=0ae882f0-ed21-4119-a08b-5c4b3f7241e7'
    # Send a GET request to the URL using the defined headers
  r = requests.get(url, headers=headers)
  r.raise_for_status()  # Raise an exception for HTTP errors
  json_data = r.json()

  # List to store product details
  products = []

  # Loop through each product retrieved from the JSON response
  for product in json_data.get("listerContent", {}).get('tiles', {}).get('content', []):
      # Extract product details
      product_name = product.get('title', '')
      product_url = f"https://www.beautybay.com{product.get('productUrl', '')}"
      product_image = product.get('images', [])[0].get('imageUrl', '') if product.get('images') else ''
      product_price = f"£{product['price']['maxValue']}" if 'price' in product and 'maxValue' in product['price'] else ''
      product_rating = product['reviews']['stats']['starRating'] if 'reviews' in product and 'stats' in product['reviews'] and 'starRating' in product['reviews']['stats'] else 'No rating found'
      total_reviews = ''
      # Append product details to the products list
      datan ={
            "Product URL": product_url,
            "Quary_Searched" : quary_name,
            "Source": source,
            "Product Name": product_name,
            "Product Rating": product_rating,
            "Total Reviews": total_reviews,
            "Product Price": product_price,
            "Product Image": product_image,

      }
      if len(products) < 10:
        products.append(datan)


  return products


def main():
    product_input = input("Enter a product name(ex. Libre Eau de Parfum): ")


    # product_input = 'Libre Eau de Parfum Women Perfume'

    product_details_faces = product_name_faces(product_input)
    product_details_sephora = product_name_sephora(product_input)
    product_details_beautybay = product_name_beautybay(product_input)
    dict_list = product_details_faces + product_details_sephora + product_details_beautybay

    df = pd.DataFrame(dict_list)

    # Save the DataFrame to an Excel file without the index
    df.to_excel('three_source_product_output.xlsx', index=False)

    # Load the workbook and select the active worksheet
    workbook = load_workbook('three_source_product_output.xlsx')
    worksheet = workbook.active

    # Iterate over columns to adjust the width
    for column_cells in worksheet.columns:
        length_list = [len(str(cell.value)) for cell in column_cells]
        max_length = max(length_list)

        # Column letter (e.g., 'A', 'B', 'C', ...)
        column_letter = get_column_letter(column_cells[0].column)

        # Adjust the column width if necessary
        worksheet.column_dimensions[column_letter].width = max_length

    # Save the workbook with adjusted column widths
    workbook.save('three_source_product_output.xlsx')


    from google.colab import files
    print('\nScraping Done!')
    files.download('three_source_product_output.xlsx')

main()
